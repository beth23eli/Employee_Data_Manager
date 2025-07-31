from flask import Blueprint, jsonify
from app.services.employeeService import EmployeeService
import openpyxl
from datetime import datetime
import os
from ..extensions import PDF
from fpdf.enums import AccessPermission, EncryptionMethod

class GeneratorController:
    def __init__(self):
        self.blueprint = Blueprint('generation', __name__)
        self._register_routes()

    def _register_routes(self):
        self.blueprint.add_url_rule('/createAggregatedEmployeeData', view_func=self.create_aggregated_employee_data, methods=['GET'])
        self.blueprint.add_url_rule('/createPdfForEmployees', view_func=self.create_pdf_for_employees, methods=['GET'])

    @staticmethod
    def create_aggregated_employee_data():
        if not os.path.exists("reports"):
            os.makedirs("reports")
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            employees = EmployeeService.get_all_employees()
            complete_data = [
                                ["ID", "NAME", "SURNAME", "CNP", "EMAIL", "POSITION", "WORKING DAYS",
                                 "TAKEN VACATION DAYS", "BONUSES", "SALARY"]
                            ] + [
                                [
                                    e.id,
                                    e.name,
                                    e.surname,
                                    e.cnp_number,
                                    e.email,
                                    str(e.position),
                                    e.num_worked_days,
                                    e.num_vacation_days,
                                    e.bonuses or 0,
                                    e.salary
                                ]
                                for e in employees
                            ]

            for row in complete_data:
                sheet.append(row)

            workbook.save("reports/employee_data_" + str(datetime.now().strftime("%B%Y")) + ".xlsx")

            return jsonify({"message": "Aggregated employee data completed successfully"}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @staticmethod
    def create_pdf_for_employees():
        employees = EmployeeService.get_all_employees()
        try:
            for employee in employees:
                pdf = PDF()
                pdf.add_page()
                pdf.set_margins(10, 10, 10)

                pdf.set_font("helvetica", size=14)
                pdf.cell(200, 10, text=f"Employee Salary Report", ln=True, align='C')
                pdf.ln(10)
                pdf.set_font("helvetica", size=12)
                pdf.cell(200, 10, text=f"Name: {employee.name} {employee.surname}", ln=True)
                pdf.cell(200, 10, text=f"CNP: {employee.cnp_number}", ln=True)
                pdf.cell(200, 10, text=f"Email: {employee.email}", ln=True)
                pdf.cell(200, 10, text=f"Position: {employee.position}", ln=True)
                pdf.cell(200, 10, text=f"Number of worked days: {employee.num_worked_days}", ln=True)
                pdf.cell(200, 10, text=f"Number of taken vacation days: {employee.num_vacation_days}", ln=True)
                pdf.cell(200, 10, text=f"Salary: {employee.salary}", ln=True)
                pdf.cell(200, 10, text=f"Bonuses: {employee.bonuses}", ln=True)

                file_name = f"reports/employee{employee.id}_" + str(datetime.now().strftime("%B%Y")) + ".pdf"
                pdf.set_encryption(
                    owner_password=employee.cnp_number,
                    user_password=employee.cnp_number,
                    encryption_method=EncryptionMethod.AES_128,
                    permissions=AccessPermission.none()
                )
                pdf.output(file_name)

            return jsonify({"message": "Personalized PDF document created for each employee successfully."}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 400

