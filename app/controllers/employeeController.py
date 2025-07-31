from flask import Blueprint, request, jsonify
from app.services.employeeService import EmployeeService
from app.services.managersService import ManagerService
import openpyxl
from datetime import datetime
from flask_mail import Message
from ..extensions import mail, PDF
from fpdf.enums import AccessPermission, EncryptionMethod
import os



class EmployeeController:
    def __init__(self):
        self.blueprint = Blueprint('employee', __name__)
        self.register_routes()

    def register_routes(self):
        self.blueprint.add_url_rule('/<int:employee_id>', view_func=self.get_employee, methods=['GET'])
        self.blueprint.add_url_rule('/create', view_func=self.create_employee, methods=['POST'])
        self.blueprint.add_url_rule('/all', view_func=self.get_all_employees, methods=['GET'])
        self.blueprint.add_url_rule('/createAggregatedEmployeeData', view_func=self.create_aggregated_employee_data, methods=['GET'])
        self.blueprint.add_url_rule('/sendAggregatedEmployeeData', view_func=self.send_aggregated_employee_data, methods=['GET'])
        self.blueprint.add_url_rule('/createPdfForEmployees', view_func=self.create_pdf_for_employees, methods=['GET'])
        self.blueprint.add_url_rule('/sendPdfToEmployees', view_func=self.send_pdf_to_employees, methods=['GET'])



    @staticmethod
    def get_employee(employee_id):
        employee = EmployeeService.get_employee(employee_id)

        if not employee:
            return jsonify({'error': 'Employee not found'}), 404
        return jsonify({
            'id': employee.id,
            'manager_id': employee.manager_id,
            'name': employee.name,
            'surname': employee.surname,
            'email': employee.email,
            'salary': employee.salary
        })

    @staticmethod
    def create_employee():
        employee_data = request.get_json()
        employee = EmployeeService.create(employee_data)

        return jsonify({'id': employee.id}), 201

    @staticmethod
    def get_all_employees():
        employees = EmployeeService.get_all_employees()

        return jsonify([
            {
                'id': employee.id,
                'name': employee.name,
                'surname': employee.surname,
                'email': employee.email,
                'manager_id': employee.manager_id
            } for employee in employees
        ])


    def do_monthly_report(self):
        try:
            self.create_aggregated_employee_data()



        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @staticmethod
    def create_aggregated_employee_data():
        if not os.path.exists("reports"):
            os.makedirs("reports")
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            employees = EmployeeService.get_all_employees()
            complete_data = [
                                ["ID", "NAME", "SURNAME", "CNP", "EMAIL", "POSITION", "WORKING DAYS",
                                 "TAKEN VACATION DAYS", "BONUSES", "SALARY"]
                            ] + [
                                [
                                    e.id,
                                    e.name,
                                    e.surname,
                                    e.cnp_number,
                                    e.email,
                                    str(e.position),
                                    e.num_worked_days,
                                    e.num_vacation_days,
                                    e.bonuses or 0,
                                    e.salary
                                ]
                                for e in employees
                            ]

            for row in complete_data:
                sheet.append(row)

            workbook.save("reports/employee_data_" + str(datetime.now().strftime("%B%Y")) + ".xlsx")

            return jsonify({"message": "Aggregated employee data completed successfully"}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @staticmethod
    def send_aggregated_employee_data():
        recipients = [manager.email for manager in ManagerService.get_all_managers()]

        try:
            msg = Message(
                subject="Monthly Employees Summary Report",
                recipients=recipients,
                body="Hello,\n\nPlease find attached the employees summary report for the last month.\n\nRegards,\nHR Team"
            )
            file_name = "employee_data_"+str(datetime.now().strftime("%B%Y"))+".xlsx"
            with open("reports/" + file_name, "rb") as f:
                msg.attach(file_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           f.read())

            mail.send(msg)

            return jsonify({"message": "Excel file sent successfully to chief manager."}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @staticmethod
    def create_pdf_for_employees():
        employees = EmployeeService.get_all_employees()
        try:
            for employee in employees:
                pdf = PDF()
                pdf.add_page()
                pdf.set_margins(10, 10, 10)

                pdf.set_font("helvetica", size=14)
                pdf.cell(200, 10, text=f"Employee Salary Report", ln=True, align='C')
                pdf.ln(10)
                pdf.set_font("helvetica", size=12)
                pdf.cell(200, 10, text=f"Name: {employee.name} {employee.surname}", ln=True)
                pdf.cell(200, 10, text=f"CNP: {employee.cnp_number}", ln=True)
                pdf.cell(200, 10, text=f"Email: {employee.email}", ln=True)
                pdf.cell(200, 10, text=f"Position: {employee.position}", ln=True)
                pdf.cell(200, 10, text=f"Number of worked days: {employee.num_worked_days}", ln=True)
                pdf.cell(200, 10, text=f"Number of taken vacation days: {employee.num_vacation_days}", ln=True)
                pdf.cell(200, 10, text=f"Salary: {employee.salary}", ln=True)
                pdf.cell(200, 10, text=f"Bonuses: {employee.bonuses}", ln=True)


                file_name = f"reports/employee{employee.id}_" + str(datetime.now().strftime("%B%Y")) + ".pdf"
                pdf.set_encryption(
                    owner_password=employee.cnp_number,
                    user_password=employee.cnp_number,
                    encryption_method=EncryptionMethod.AES_128,
                    permissions=AccessPermission.none()
                )
                pdf.output(file_name)

            return jsonify({"message": "Personalized PDF document created for each employee successfully."}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @staticmethod
    def send_pdf_to_employees():
        employees = EmployeeService.get_all_employees()
        employee = employees[0]

        # for employee in employees:

        try:
            msg = Message(
                subject="Payroll Report – " + str(datetime.now().strftime("%B %Y")),
                recipients=["secrieru.eliza@gmail.com"],
                body="Attached to this email, you will find your salary report for the month of "+ str(datetime.now().strftime("%B%Y")) + ".\nTo open the file, the password is your personal identification number - CNP.\n\nThank you!\nPayroll Team"
            )
            file_name = "employee" + str(employee.id) + "_"+ str(datetime.now().strftime("%B%Y")) + ".pdf"
            with open("reports/" + file_name, "rb") as f:
                msg.attach(file_name, "application/pdf", f.read())

            mail.send(msg)

            return jsonify({f"message": f"Pdf file sent successfully to employee {employee.id}"}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 400
