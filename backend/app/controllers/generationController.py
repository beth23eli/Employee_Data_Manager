from flask import Blueprint, jsonify
from ..services.employeeService import EmployeeService
import openpyxl
import os
from ..extensions import PDF, previous_month
from fpdf.enums import AccessPermission, EncryptionMethod
import zipfile


class GeneratorController:
    is_completed = False

    def __init__(self):
        self.is_completed = False
        self.blueprint = Blueprint('generation', __name__)
        self._register_routes()

    def _register_routes(self):
        self.blueprint.add_url_rule('/createAggregatedEmployeeData', view_func=self.create_aggregated_employee_data, methods=['POST'])
        self.blueprint.add_url_rule('/createPdfForEmployees', view_func=self.create_pdf_for_employees, methods=['POST'])
        self.blueprint.add_url_rule('/createArchive', view_func=self.create_archive, methods=['POST'])


    def create_aggregated_employee_data(self):
        if not os.path.exists("reports"):
            os.makedirs("reports")
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            employees = EmployeeService.get_all_employees()
            complete_data = [
                                ["ID", "NAME", "SURNAME", "CNP", "EMAIL", "POSITION", "WORKING DAYS",
                                 "TAKEN VACATION DAYS", "BONUSES", "SALARY"]
                            ] + [
                                [
                                    e.id,
                                    e.name,
                                    e.surname,
                                    e.cnp_number,
                                    e.email,
                                    str(e.position.title),
                                    e.num_worked_days,
                                    e.num_vacation_days,
                                    e.bonuses or 0,
                                    e.position.salary
                                ]
                                for e in employees
                            ]

            for row in complete_data:
                sheet.append(row)

            workbook.save("reports/employee_data_" + str(previous_month.strftime("%B%Y")) + ".xlsx")

            self.is_completed = True
            return jsonify({"message": "Aggregated employee data completed successfully"}), 200
        except Exception as e:
            self.is_completed = False
            return jsonify({"error": str(e)}), 400


    def create_pdf_for_employees(self):
        if self.is_completed:
            employees = EmployeeService.get_all_employees()
            try:
                for employee in employees:
                    pdf = PDF()
                    pdf.add_page()
                    pdf.set_margins(10, 10, 10)

                    pdf.set_font("helvetica", size=14)
                    pdf.cell(200, 10, text=f"Employee Salary Report", ln=True, align='C')
                    pdf.ln(10)
                    pdf.set_font("helvetica", size=12)
                    pdf.cell(200, 10, text=f"Name: {employee.name} {employee.surname}", ln=True)
                    pdf.cell(200, 10, text=f"CNP: {employee.cnp_number}", ln=True)
                    pdf.cell(200, 10, text=f"Email: {employee.email}", ln=True)
                    pdf.cell(200, 10, text=f"Position: {employee.position.title}", ln=True)
                    pdf.cell(200, 10, text=f"Number of worked days: {employee.num_worked_days}", ln=True)
                    pdf.cell(200, 10, text=f"Number of taken vacation days: {employee.num_vacation_days}", ln=True)
                    pdf.cell(200, 10, text=f"Salary: {employee.position.salary}", ln=True)
                    pdf.cell(200, 10, text=f"Bonuses: {employee.bonuses}", ln=True)

                    file_name = f"reports/employee{employee.id}_" + str(previous_month.strftime("%B%Y")) + ".pdf"
                    pdf.set_encryption(
                        owner_password=employee.cnp_number,
                        user_password=employee.cnp_number,
                        encryption_method=EncryptionMethod.AES_128,
                        permissions=AccessPermission.none()
                    )
                    pdf.output(file_name)

                self.is_completed = True
                return jsonify({"message": "Personalized PDF document created for each employee successfully."}), 200
            except Exception as e:
                self.is_completed = False
                return jsonify({"error": str(e)}), 400

        return jsonify({"message": "It seems like the aggregated employee data was not created."}), 400

    @staticmethod
    def create_archive():
        path = 'reports/'
        files = os.listdir(path)

        if not os.path.exists("archives"):
            os.makedirs("archives")

        try:
            with zipfile.ZipFile("archives/monthly_report_" + str(previous_month.strftime("%B%Y")) + ".zip", mode="w") as archive:
                for file in files:
                    full_path = os.path.join(path, file)
                    archive.write(full_path, arcname=file)

            return jsonify({"message": "Archive created successfully."})
        except Exception as e:
            return jsonify({"error": str(e)})